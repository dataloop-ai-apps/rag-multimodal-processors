"""
Excel (XLS/XLSX) extraction logic.

Handles Excel-specific extraction operations:
- Text extraction from cells across all sheets
- Table extraction with markdown conversion
- Image extraction from embedded resources
"""

import logging
import os
import tempfile
from typing import List, Dict, Tuple, Optional, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io

from utils.extracted_data import ExtractedData
from utils.data_types import ImageContent, TableContent

logger = logging.getLogger("rag-preprocessor")


class XLSExtractor:
    """Excel extraction operations."""

    @staticmethod
    def extract(data: ExtractedData) -> ExtractedData:
        """Extract content from Excel item."""
        data.current_stage = "extraction"

        if not data.item:
            data.log_error("No item provided for extraction")
            return data

        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                file_path = data.item.download(local_path=temp_dir)
                use_markdown = data.config.use_markdown_extraction

                # Extract images if configured
                if data.config.extract_images:
                    data.images = XLSExtractor._extract_images(file_path, temp_dir)

                # Extract tables if configured
                if data.config.extract_tables:
                    data.tables = XLSExtractor._extract_tables(file_path)

                # Extract content based on use_markdown_extraction setting
                if use_markdown:
                    data.content_text = XLSExtractor._extract_markdown(file_path, data.tables)
                else:
                    data.content_text = XLSExtractor._extract_plain_text(file_path)

                # Set metadata
                data.metadata = {
                    'source_file': data.item_name,
                    'extraction_method': 'pandas-openpyxl',
                    'format': 'markdown' if use_markdown else 'plain',
                    'image_count': len(data.images),
                    'table_count': len(data.tables),
                    'processor': 'xls',
                }

        except Exception as e:
            data.log_error("Excel extraction failed. Check logs for details.")
            logger.exception(f"Excel extraction error: {e}")

        return data

    @staticmethod
    def _extract_plain_text(file_path: str) -> str:
        """Extract plain text from Excel without formatting."""
        text_parts = []
        
        try:
            # Read all sheets
            with pd.ExcelFile(file_path, engine='openpyxl') as excel_file:
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    # Add sheet header
                    text_parts.append(f"\n\n--- Sheet: {sheet_name} ---\n\n")
                    
                    # Convert DataFrame to text
                    for idx, row in df.iterrows():
                        row_values = [str(val) if pd.notna(val) else '' for val in row]
                        row_text = ' | '.join(row_values)
                        if row_text.strip():
                            text_parts.append(row_text)
                
        except Exception as e:
            logger.warning(f"Error extracting plain text from Excel: {e}")
            return ""

        return '\n'.join(text_parts)

    @staticmethod
    def _extract_markdown(file_path: str, tables: List[TableContent]) -> str:
        """Convert Excel content to markdown format."""
        md_parts = []
        table_iter = iter(tables)
        current_table = next(table_iter, None)

        try:
            with pd.ExcelFile(file_path, engine='openpyxl') as excel_file:
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    # Add sheet header
                    md_parts.append(f"\n\n## Sheet: {sheet_name}\n\n")
                    
                    # Convert DataFrame to markdown table
                    if not df.empty:
                        # Use first row as header if it looks like headers
                        if XLSExtractor._looks_like_header(df.iloc[0]):
                            headers = [str(val) if pd.notna(val) else '' for val in df.iloc[0]]
                            data_rows = df.iloc[1:]
                        else:
                            headers = [f"Column {i+1}" for i in range(len(df.columns))]
                            data_rows = df
                        
                        # Create markdown table
                        md_table = XLSExtractor._dataframe_to_markdown(headers, data_rows)
                        if md_table:
                            md_parts.append(md_table)
                    
                    # Insert extracted table if available
                    if current_table:
                        md_parts.append(current_table.markdown)
                        current_table = next(table_iter, None)

        except Exception as e:
            logger.warning(f"Error extracting markdown from Excel: {e}")

        return '\n\n'.join(md_parts)

    @staticmethod
    def _looks_like_header(row: pd.Series) -> bool:
        """Check if a row looks like a header row."""
        # Simple heuristic: if most values are non-numeric strings, it's likely a header
        non_numeric_count = 0
        for val in row:
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str and not val_str.replace('.', '').replace('-', '').isdigit():
                    non_numeric_count += 1
        
        return non_numeric_count > len(row) * 0.5

    @staticmethod
    def _dataframe_to_markdown(headers: List[str], data_rows: pd.DataFrame) -> str:
        """Convert DataFrame rows to markdown table."""
        if not headers:
            return ""

        md = "| " + " | ".join(headers) + " |\n"
        md += "| " + " | ".join(["---"] * len(headers)) + " |\n"

        for _, row in data_rows.iterrows():
            values = [str(val) if pd.notna(val) else '' for val in row]
            # Pad with empty strings if row is shorter than headers
            while len(values) < len(headers):
                values.append('')
            # Truncate if row is longer than headers
            values = values[:len(headers)]
            md += "| " + " | ".join(values) + " |\n"

        return md

    @staticmethod
    def _extract_images(file_path: str, temp_dir: str) -> List[ImageContent]:
        """Extract embedded images from Excel with size metadata."""
        images = []

        try:
            workbook = load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract images from worksheet
                # openpyxl stores images in sheet._images list
                if hasattr(sheet, '_images') and sheet._images:
                    for img_index, img in enumerate(sheet._images):
                        try:
                            # Get image data - openpyxl Image objects have _data() method
                            if hasattr(img, '_data'):
                                image_data = img._data()
                            elif hasattr(img, 'ref'):
                                # Alternative: try to get from anchor if available
                                logger.warning(f"Image {img_index} in sheet {sheet_name} has no _data() method")
                                continue
                            else:
                                continue
                            
                            # Determine file extension
                            ext = 'png'  # Default
                            if hasattr(img, 'format') and img.format:
                                ext = img.format.lower()
                                if ext not in ['png', 'jpeg', 'jpg', 'gif']:
                                    ext = 'png'
                            
                            filename = f"sheet_{sheet_name}_img{img_index}.{ext}"
                            # Sanitize filename
                            filename = filename.replace('/', '_').replace('\\', '_')
                            image_path = os.path.join(temp_dir, filename)
                            
                            # Save image
                            with open(image_path, 'wb') as f:
                                f.write(image_data)
                            
                            # Extract image dimensions using PIL
                            size = XLSExtractor._get_image_size(image_data)
                            
                            images.append(
                                ImageContent(
                                    path=image_path,
                                    format=ext,
                                    size=size,
                                    caption=f"Image from sheet: {sheet_name}"
                                )
                            )
                        except (IOError, OSError, ValueError, KeyError, AttributeError) as e:
                            logger.warning(f"Failed to extract image {img_index} from sheet {sheet_name}: {e}")

        except Exception as e:
            logger.warning(f"Error extracting images from Excel: {e}")

        return images

    @staticmethod
    def _get_image_size(image_data: bytes) -> Optional[Tuple[int, int]]:
        """Extract image dimensions from blob data."""
        try:
            with Image.open(io.BytesIO(image_data)) as img:
                return img.size
        except (IOError, OSError, ValueError):
            return None

    @staticmethod
    def _extract_tables(file_path: str) -> List[TableContent]:
        """Extract tables from Excel with markdown conversion."""
        tables = []

        try:
            with pd.ExcelFile(file_path, engine='openpyxl') as excel_file:
                for sheet_index, sheet_name in enumerate(excel_file.sheet_names):
                    try:
                        # Read without headers first to check if first row is actually a header
                        df_no_header = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                        
                        if df_no_header.empty:
                            continue
                        
                        # Check if first row looks like a header
                        if XLSExtractor._looks_like_header(df_no_header.iloc[0]):
                            # First row is header, use it and skip it from data
                            headers = [str(val) if pd.notna(val) else f"Column {i+1}" 
                                      for i, val in enumerate(df_no_header.iloc[0])]
                            data_df = df_no_header.iloc[1:].reset_index(drop=True)
                            # Set column names to indices for easier access
                            data_df.columns = range(len(headers))
                        else:
                            # First row is data, use generic column names
                            headers = [f"Column {i+1}" for i in range(len(df_no_header.columns))]
                            data_df = df_no_header.reset_index(drop=True)
                            data_df.columns = range(len(headers))
                        
                        # Convert to list of dictionaries
                        rows = []
                        for _, row in data_df.iterrows():
                            row_data = {}
                            for i, header in enumerate(headers):
                                if i < len(row):
                                    val = row.iloc[i]
                                    row_data[header] = str(val) if pd.notna(val) else ''
                                else:
                                    row_data[header] = ''
                            rows.append(row_data)
                        
                        # Generate markdown
                        markdown = XLSExtractor._table_to_markdown(headers, rows)
                        
                        tables.append(
                            TableContent(
                                data=rows,
                                markdown=markdown,
                                page_number=sheet_index + 1
                            )
                        )
                    except (ValueError, AttributeError, IndexError) as e:
                        logger.warning(f"Failed to extract table from sheet {sheet_name}: {e}")

        except Exception as e:
            logger.warning(f"Error extracting tables from Excel: {e}")

        return tables

    @staticmethod
    def _table_to_markdown(headers: List[str], rows: List[Dict[str, str]]) -> str:
        """Convert table data to markdown format."""
        if not headers:
            return ""

        md = "| " + " | ".join(headers) + " |\n"
        md += "| " + " | ".join(["---"] * len(headers)) + " |\n"

        for row in rows:
            values = [str(row.get(h, '')) for h in headers]
            md += "| " + " | ".join(values) + " |\n"

        return md
